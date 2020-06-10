import openpyxl
import FileSearch
import Filehash
from java import Run


def file_Path_Create_List(target_folder, target_file_type):
    return {
        "File_All_List": FileSearch.get_file_path(
            file_path=target_folder, target_file=target_file_type),

        "File_Path_List": FileSearch.get_file_directory(file_path=target_folder),

        "File_Name_List": FileSearch.get_file_name(
            file_path=target_folder, target_file=target_file_type)
    }


def createExcelFile(menberData_path, TestFile_NameList, outputfilename):

    wb = openpyxl.load_workbook('temp.xlsx')
    sheet = wb[wb.sheetnames[0]]

    # 範囲を指定してセルを取得する
    cells = sheet['B3':'G2000']

    with open(menberData_path, 'r', encoding='utf-8') as f:
        for menberCount in [n * len(TestFile_NameList) for n in list(range(sum([1 for _ in open(menberData_path)])))]:
            for testFilename in TestFile_NameList:
                if testFilename == TestFile_NameList[0]:
                    cells[menberCount][0].value = f.readline().replace("\n", "")
                    cells[menberCount][1].value = testFilename
                else:
                    cells[menberCount +
                          TestFile_NameList.index(testFilename)][1].value = testFilename
            sheet.merge_cells('B{}:B{}'.format(
                menberCount + 3, menberCount + 3 + len(TestFile_NameList) - 1))
            sheet.merge_cells('G{}:G{}'.format(
                menberCount + 3, menberCount + 3 + len(TestFile_NameList) - 1))

        # ワークブックに名前をつけて保存する
        wb.save(outputfilename + '.xlsx')


def openExcellFile(saitenpath, output_FileName):
    wb = openpyxl.load_workbook(output_FileName + '.xlsx')
    sheet = wb[wb.sheetnames[0]]

    # 範囲を指定してセルを取得する
    cells = sheet['B3':'G2000']

    schoolID = ""
    student = saitenpath + "\\{}\\"
    File = {}
    for row in cells:
        if row[0].value is not None:
            schoolID = row[0].value
            File = file_Path_Create_List(
                student.format(schoolID), "**/?*")
        else:
            File = file_Path_Create_List(
                student.format(schoolID), "**/?*")

        # 提出されていたら〇
        if row[1].value in File["File_Name_List"]:
            row[2].value = "〇"
            # コピーの可能性
            row[3].value = Filehash.File_hash(
                File["File_All_List"][File["File_Name_List"].index(row[1].value)])

            # javaコンパイル
            if ".java" in row[1].value:
                for javaFile_path in [s for s in File["File_All_List"] if s.endswith('.java')]:
                    Run.java_compile(javaFile_path)
                # コンパイルファイル検索
                File = file_Path_Create_List(
                    student.format(schoolID), "**/?*.class")
                if row[1].value[:-5] + ".class" in File["File_Name_List"]:
                    row[4].value = "〇"

    wb.save(output_FileName + '.xlsx')


if __name__ == "__main__":
    # 対象学生の名簿
    menberList = "C:\\Users\\tsukasa\\Documents\\学校\\副手\\saiten\\unberlist.txt"
    # 課題ファイル名のリスト作成
    TestFile_List = ["Person.java",
                     "SampleYen.java",
                     "Yen.java"]
    student = "C:\\Users\\tsukasa\\Documents\\学校\\副手\\18FI009\\"

    with open(menberList, 'r', encoding='utf-8') as f:
        createExcelFile(menberList, TestFile_List)
        openExcellFile()
